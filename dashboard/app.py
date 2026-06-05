"""
Genpact SAP Migration Validator — Dashboard V4
Run:  python dashboard/app.py
Open: http://localhost:5000
"""

import sys
import csv as csv_mod
import threading
import time
import json
import io
from pathlib import Path
from datetime import datetime

from flask import Flask, render_template, jsonify, send_file, request, Response
from werkzeug.utils import secure_filename

sys.path.insert(0, str(Path(__file__).parent.parent))

from core.validator    import MaterialValidator
from core.reporter     import generate_excel_report
from core.field_labels import (
    get_label, load_custom_labels, enrich_field_rows,
    get_display, SAP_FIELD_LABELS
)
from core.field_mapper  import build_field_mapping, mapping_result_to_dict
from core.object_config import get_object_config, SAP_OBJECT_CONFIG

app = Flask(__name__)

BASE_DIR      = Path(__file__).parent.parent
REPORTS_DIR   = BASE_DIR / "reports"
CONFIG_FILE   = BASE_DIR / "config.json"
LABELS_FILE   = BASE_DIR / "custom_labels.csv"
TEMPLATES_DIR = BASE_DIR / "templates"

REPORTS_DIR.mkdir(parents=True, exist_ok=True)
TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)

DEFAULT_CONFIG = {
    "source_dir":      str(BASE_DIR / "data" / "source"),
    "target_dir":      str(BASE_DIR / "data" / "target"),
    "pass_threshold":  100.0,
    "selected_fields": [],
    "manual_pairs":    [],
    "active_template": "",
}

results_store = {}
scan_status   = {
    "last_scan": None, "scanning": False, "error": None,
    "current_file": None, "total_files": 0, "completed_files": 0,
}
file_states  = {}
activity_log = []

SUPPORTED_EXT = {".csv", ".xlsx", ".xls"}
TEMPLATE_EXT  = {".csv", ".xlsx", ".xls", ".txt"}
scan_lock     = threading.Lock()


# ── Config helpers ──────────────────────────────────────────────────────────────

def load_config():
    if CONFIG_FILE.exists():
        try:
            return {**DEFAULT_CONFIG, **json.loads(CONFIG_FILE.read_text())}
        except Exception as e:
            print(f"Config load failed: {e}")
    return dict(DEFAULT_CONFIG)


def save_config(cfg):
    CONFIG_FILE.write_text(json.dumps(cfg, indent=2))


def get_dirs():
    cfg = load_config()
    src = Path(cfg.get("source_dir", DEFAULT_CONFIG["source_dir"]))
    tgt = Path(cfg.get("target_dir", DEFAULT_CONFIG["target_dir"]))
    src.mkdir(parents=True, exist_ok=True)
    tgt.mkdir(parents=True, exist_ok=True)
    return src, tgt


def log_event(message, level="info"):
    entry = {"ts": datetime.now().strftime("%H:%M:%S"), "message": message, "level": level}
    activity_log.append(entry)
    if len(activity_log) > 200:
        activity_log.pop(0)
    print(f"  [{entry['ts']}] [{level.upper()}] {message}")


def cleanup_old_reports(keep=20):
    files = sorted(REPORTS_DIR.glob("*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
    for f in files[keep:]:
        try: f.unlink()
        except: pass


def _get_custom_labels():
    return load_custom_labels(str(LABELS_FILE)) if LABELS_FILE.exists() else {}


def _read_file_headers(src_path: str, tgt_path: str = None):
    """Read only the header row — very fast, no full file load."""
    def headers(path):
        p = Path(path)
        if not p.exists():
            return []
        if p.suffix.lower() in (".xlsx", ".xls"):
            import openpyxl
            wb   = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
            ws   = wb.active
            cols = [str(c.value).strip().upper()
                    for c in next(ws.iter_rows(max_row=1)) if c.value]
            wb.close()
            return cols
        with open(str(p), encoding="utf-8-sig") as f:
            reader = csv_mod.reader(f)
            return [c.strip().upper() for c in next(reader)]

    src_cols = headers(src_path) if src_path else []
    tgt_cols = headers(tgt_path) if tgt_path else []
    return src_cols, tgt_cols


# ── Field-selection template helpers ───────────────────────────────────────────

def _read_template_fields(path: Path) -> list:
    """
    Parse a field-selection template file.
    Supported formats:
      CSV  : first column = field names (header row auto-skipped)
      XLSX : column A = field names (header row auto-skipped)
      TXT  : one field name per line (lines starting with # are comments)
    Returns: list of UPPERCASE field names, no blanks, no comments.
    """
    fields = []
    suffix = path.suffix.lower()
    try:
        if suffix == ".txt":
            for line in path.read_text(encoding="utf-8-sig").splitlines():
                val = line.strip()
                if val and not val.startswith("#"):
                    fields.append(val.upper())

        elif suffix in (".xlsx", ".xls"):
            import openpyxl
            wb        = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            ws        = wb.active
            first_row = True
            for row in ws.iter_rows(values_only=True):
                val = str(row[0] or "").strip()
                if not val:
                    continue
                if val.startswith("#"):
                    continue
                val_up = val.upper()
                if first_row:
                    first_row = False
                    if val_up in ("FIELD", "FIELD_NAME", "FIELDS", "SAP_FIELD",
                                  "FIELDNAME", "SAP FIELD", "FIELD NAME"):
                        continue
                fields.append(val_up)
            wb.close()

        else:  # CSV
            with open(str(path), encoding="utf-8-sig") as f:
                reader    = csv_mod.reader(f)
                first_row = True
                for row in reader:
                    if not row:
                        continue
                    val = row[0].strip()
                    if not val:
                        continue
                    # Skip comment lines
                    if val.startswith("#"):
                        continue
                    val_up = val.upper()
                    # Skip header row
                    if first_row:
                        first_row = False
                        if val_up in ("FIELD", "FIELD_NAME", "FIELDS", "SAP_FIELD",
                                      "FIELDNAME", "SAP FIELD", "FIELD NAME"):
                            continue
                    fields.append(val_up)

    except Exception as e:
        print(f"Template parse error ({path.name}): {e}")
    return fields


def _list_templates() -> list:
    cfg    = load_config()
    active = cfg.get("active_template", "")
    result = []
    for p in sorted(TEMPLATES_DIR.iterdir()):
        if p.suffix.lower() in TEMPLATE_EXT and p.is_file():
            fields = _read_template_fields(p)
            result.append({
                "filename":    p.name,
                "field_count": len(fields),
                "fields":      fields,
                "is_active":   p.name == active,
                "modified":    datetime.fromtimestamp(
                    p.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
            })
    return result


# ── File discovery & pairing ────────────────────────────────────────────────────

def get_available_files():
    src_dir, tgt_dir = get_dirs()
    src = sorted([f for f in src_dir.iterdir() if f.suffix.lower() in SUPPORTED_EXT],
                 key=lambda f: f.name.upper())
    tgt = sorted([f for f in tgt_dir.iterdir() if f.suffix.lower() in SUPPORTED_EXT],
                 key=lambda f: f.name.upper())
    return src, tgt


def discover_pairs():
    SOURCE_DIR, TARGET_DIR = get_dirs()
    cfg          = load_config()
    manual_pairs = cfg.get("manual_pairs", [])

    src_files = {f.name: f for f in SOURCE_DIR.iterdir()
                 if f.suffix.lower() in SUPPORTED_EXT}
    tgt_files = {f.name: f for f in TARGET_DIR.iterdir()
                 if f.suffix.lower() in SUPPORTED_EXT}

    pairs    = []
    used_src = set()
    used_tgt = set()

    # Manual pairs first
    for mp in manual_pairs:
        src_name = mp.get("source_file", "")
        tgt_name = mp.get("target_file", "")
        name     = mp.get("name", "").upper().strip() or Path(src_name).stem.upper()
        sp       = str(src_files[src_name]) if src_name in src_files else None
        tp       = str(tgt_files[tgt_name]) if tgt_name in tgt_files else None
        has_pair = bool(sp and tp)
        mtime    = max(Path(sp).stat().st_mtime, Path(tp).stat().st_mtime) if has_pair else None
        pairs.append({
            "name": name, "source_path": sp, "target_path": tp,
            "has_pair": has_pair, "mtime": mtime,
            "source_file": Path(sp).name if sp else src_name,
            "target_file": Path(tp).name if tp else tgt_name,
            "match_type": "manual",
            "missing": [] if has_pair else
                       (["source"] if not sp else []) + (["target"] if not tp else []),
        })
        if sp: used_src.add(src_name)
        if tp: used_tgt.add(tgt_name)

    # Auto-pair by exact filename stem
    src_by_stem = {Path(f).stem.upper(): (f, fp)
                   for f, fp in src_files.items() if f not in used_src}
    tgt_by_stem = {Path(f).stem.upper(): (f, fp)
                   for f, fp in tgt_files.items() if f not in used_tgt}

    for stem in sorted(set(src_by_stem) & set(tgt_by_stem)):
        sf, sp = src_by_stem[stem]
        tf, tp = tgt_by_stem[stem]
        mtime  = max(sp.stat().st_mtime, tp.stat().st_mtime)
        pairs.append({
            "name": stem, "source_path": str(sp), "target_path": str(tp),
            "has_pair": True, "mtime": mtime,
            "source_file": sf, "target_file": tf,
            "match_type": "auto", "missing": [],
        })
        used_src.add(sf)
        used_tgt.add(tf)

    # Unmatched
    for f, fp in src_files.items():
        if f not in used_src:
            pairs.append({"name": Path(f).stem.upper(), "source_path": str(fp),
                          "target_path": None, "has_pair": False, "mtime": None,
                          "source_file": f, "target_file": None,
                          "match_type": "unmatched", "missing": ["target"]})
    for f, fp in tgt_files.items():
        if f not in used_tgt:
            pairs.append({"name": Path(f).stem.upper(), "source_path": None,
                          "target_path": str(fp), "has_pair": False, "mtime": None,
                          "source_file": None, "target_file": f,
                          "match_type": "unmatched", "missing": ["source"]})
    return pairs


# ── Business status ─────────────────────────────────────────────────────────────

def calculate_business_status(result, pass_threshold):
    ss        = result.summary_stats
    pass_rate = float(ss.get("pass_rate_pct", 0))
    only_src  = int(result.records_only_in_source or 0)
    only_tgt  = int(result.records_only_in_target or 0)

    if pass_rate < pass_threshold:
        return {"status": "FAIL", "field_status": "FAIL", "record_status": "CHECKED",
                "message": (f"Field validation failed. Pass rate is {pass_rate:.2f}% "
                            f"which is below threshold {pass_threshold:.2f}%.")}

    if only_src > 0 or only_tgt > 0:
        if only_src > 0 and only_tgt > 0:
            msg = (f"{only_src:,} records only in source and "
                   f"{only_tgt:,} records only in target.")
        elif only_tgt > 0:
            msg = f"Target has {only_tgt:,} extra records not in source."
        else:
            msg = f"Source has {only_src:,} records not found in target."
        return {"status": "WARNING", "field_status": "PASS", "record_status": "WARNING",
                "message": (f"Field validation passed ({pass_rate:.2f}% >= "
                            f"{pass_threshold:.2f}%), but {msg}")}

    return {"status": "PASS", "field_status": "PASS", "record_status": "PASS",
            "message": (f"Validation passed. Field pass rate {pass_rate:.2f}% "
                        f"and records fully reconciled.")}


# ── Core validation runner ──────────────────────────────────────────────────────

def run_validation(name, source_path, target_path):
    cfg            = load_config()
    pass_threshold = float(cfg.get("pass_threshold", 100.0))
    custom         = _get_custom_labels()

    # ── Determine field filter: template > manual > all ───────────────────────
    selected_fields    = cfg.get("selected_fields", [])
    active_template    = cfg.get("active_template", "")
    template_name_used = ""

    if active_template:
        tmpl_path = TEMPLATES_DIR / active_template
        if tmpl_path.exists():
            tmpl_fields = _read_template_fields(tmpl_path)
            if tmpl_fields:
                selected_fields    = tmpl_fields
                template_name_used = active_template
                log_event(
                    f"{name}: using template '{active_template}' "
                    f"({len(tmpl_fields)} fields)",
                    "info",
                )
            else:
                log_event(
                    f"{name}: template '{active_template}' is empty — "
                    f"validating all fields",
                    "warn",
                )
        else:
            log_event(
                f"{name}: template '{active_template}' not found — "
                f"validating all fields",
                "warn",
            )

    obj_cfg  = get_object_config(name)
    join_key = obj_cfg.get("join_key", None)

    src_mb = Path(source_path).stat().st_size / (1024 * 1024)
    tgt_mb = Path(target_path).stat().st_size / (1024 * 1024)
    if src_mb > 50 or tgt_mb > 50:
        log_event(
            f"{name}: large files ({src_mb:.1f} MB / {tgt_mb:.1f} MB) — "
            f"may take a few minutes",
            "warn",
        )

    try:
        src_cols, tgt_cols = _read_file_headers(source_path, target_path)
    except Exception as e:
        log_event(f"{name}: could not read headers — {e}", "warn")
        src_cols, tgt_cols = [], []

    jk_upper  = join_key.upper() if join_key else ""
    src_no_jk = [c for c in src_cols if c != jk_upper]
    tgt_no_jk = [c for c in tgt_cols if c != jk_upper]

    # ── Build field mapping: source → target ──────────────────────────────────
    # Forward approach: for each source column find its target equivalent.
    # The target file drives WHAT gets validated — we use all target columns
    # as the reference set. selected_fields/template filters which target
    # columns we care about.

    if selected_fields:
        sel_upper = set(s.upper() for s in selected_fields)
        # Filter to target columns the user selected
        tgt_filtered = [c for c in tgt_no_jk if c in sel_upper]
        if not tgt_filtered:
            tgt_filtered = tgt_no_jk  # fallback: all target columns
        log_event(
            f"{name}: filtering to {len(tgt_filtered)} of "
            f"{len(tgt_no_jk)} target columns",
            "info",
        )
    else:
        tgt_filtered = tgt_no_jk

    # Step 1: forward mapping — finds source col for each source field,
    # matched against target set (exact + alias + fuzzy)
    mapping_result = build_field_mapping(
        source_cols=src_no_jk,
        target_cols=tgt_filtered,   # only validate fields that exist in target
        object_type=name,
        selected_fields=None,       # filtering already done above via tgt_filtered
        custom_labels=custom,
    )

    field_map    = mapping_result.mapped_fields
    exact_count  = sum(1 for d in mapping_result.mapped_details if d.method == "exact")
    alias_count  = sum(1 for d in mapping_result.mapped_details if "alias" in d.method)
    fuzzy_count  = sum(1 for d in mapping_result.mapped_details if d.method == "fuzzy")

    log_event(
        f"{name}: mapped {len(field_map)} fields "
        f"({exact_count} exact, {alias_count} alias, {fuzzy_count} fuzzy)"
        + (f" via template '{template_name_used}'" if template_name_used else ""),
        "info",
    )

    # Warn if template produced zero mapped fields
    if selected_fields and template_name_used and not field_map:
        log_event(
            f"{name}: WARNING — template '{template_name_used}' has "
            f"{len(selected_fields)} fields but NONE matched any source column. "
            f"Source columns: {src_no_jk[:8]}... "
            f"Template fields: {selected_fields[:8]}...",
            "error",
        )
    elif selected_fields and template_name_used:
        # Log which template fields weren't found
        mapped_set = set(field_map.keys())
        src_set    = set(src_no_jk)
        missing    = [f for f in selected_fields
                      if f not in src_set and f not in mapped_set]
        if missing:
            log_event(
                f"{name}: {len(missing)} template field(s) not matched: "
                f"{', '.join(missing[:8])}"
                + (" …" if len(missing) > 8 else ""),
                "warn",
            )

    validator = MaterialValidator(
        field_map=field_map,
        pass_threshold=pass_threshold,
        join_key=join_key,
        custom_labels=custom if custom else None,
    )

    result          = validator.validate(source_path, target_path)
    ss              = result.summary_stats
    business_status = calculate_business_status(result, pass_threshold)

    # ── Build field rows ───────────────────────────────────────────────────────
    field_rows = []
    for fr in result.field_results:
        detail = next(
            (d for d in mapping_result.mapped_details
             if d.source_field == fr.field_source), None
        )
        disp = get_display(fr.field_source, fr.field_target, custom)
        field_rows.append({
            "field":              fr.field_source,
            "field_label":        disp["source_label"],
            "field_target":       fr.field_target,
            "field_target_label": disp["target_label"],
            "display_name":       disp["display_name"],
            "display_mapping":    disp["display_mapping"],
            "is_cross_mapped":    disp["is_cross_mapped"],
            "mapping_method":     detail.method if detail else "exact",
            "mapping_confidence": detail.confidence if detail else 1.0,
            "type":               "numeric" if fr.is_numeric else "string",
            "tolerance":          fr.tolerance_used,
            "total":              fr.total_records,
            "matched":            fr.matched,
            "mismatched":         fr.mismatched,
            "miss_source":        fr.missing_in_source,
            "miss_target":        fr.missing_in_target,
            "match_pct":          fr.match_pct,
            "pass_threshold":     fr.pass_threshold,
            "status":             fr.status,
            "mismatches":         fr.mismatch_details,
            "mismatch_count":     len(fr.mismatch_details),
        })

    # ── Mapping info for dashboard ─────────────────────────────────────────────
    mapping_info = None
    if result.mapping:
        mapping_info = {
            "join_key":           result.mapping.join_key,
            "join_key_label":     get_label(result.mapping.join_key, custom),
            "matched_fields":     result.mapping.matched_fields,
            "matched_labels":     {f: get_label(f, custom)
                                   for f in result.mapping.matched_fields},
            "source_only_fields": mapping_result.unmapped_source,
            "source_only_labels": {f: get_label(f, custom)
                                   for f in mapping_result.unmapped_source},
            "target_only_fields": mapping_result.unmapped_target,
            "target_only_labels": {f: get_label(f, custom)
                                   for f in mapping_result.unmapped_target},
            "numeric_fields":     result.mapping.numeric_fields,
            "tolerance_map":      result.mapping.tolerance_map,
            "selected_fields":    selected_fields,
            "pass_threshold":     pass_threshold,
        }

    # ── available_fields for Settings field selector ───────────────────────────
    sel_set = set(selected_fields) if selected_fields else set()
    available_fields = []
    for col in src_cols:
        tgt_col = field_map.get(col)
        available_fields.append({
            "field":        col,
            "label":        get_label(col, custom),
            "in_source":    True,
            "in_target":    tgt_col is not None,
            "target_col":   tgt_col or "",
            "target_label": get_label(tgt_col, custom) if tgt_col else "",
            "common":       tgt_col is not None,
            "selected":     not sel_set or col in sel_set,
        })
    for col in tgt_cols:
        if col not in field_map.values() and col != jk_upper:
            available_fields.append({
                "field":        col,
                "label":        get_label(col, custom),
                "in_source":    False,
                "in_target":    True,
                "target_col":   col,
                "target_label": get_label(col, custom),
                "common":       False,
                "selected":     False,
            })

    ts             = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"{name}_{ts}.xlsx"
    excel_path     = REPORTS_DIR / excel_filename

    result_dict = {
        "name":                   name,
        "sap_object":             obj_cfg.get("description", name),
        "status":                 business_status["status"],
        "validator_status":       result.overall_status,
        "field_status":           business_status["field_status"],
        "record_status":          business_status["record_status"],
        "business_message":       business_status["message"],
        "source_file":            Path(source_path).name,
        "target_file":            Path(target_path).name,
        "total_source_records":   result.total_source_records,
        "total_target_records":   result.total_target_records,
        "records_matched":        result.records_matched,
        "records_only_in_source": result.records_only_in_source,
        "records_only_in_target": result.records_only_in_target,
        "fields_passed":          ss["fields_passed"],
        "fields_failed":          ss["fields_failed"],
        "total_fields":           ss["total_fields_validated"],
        "pass_rate_pct":          ss["pass_rate_pct"],
        "pass_threshold":         pass_threshold,
        "selected_fields":        selected_fields,
        "template_used":          template_name_used,
        "errors":                 result.errors,
        "mapping":                mapping_info,
        "field_mapping_detail":   mapping_result_to_dict(mapping_result),
        "field_results":          field_rows,
        "available_fields":       available_fields,
        "run_at":                 datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "excel_file":             excel_filename,
    }

    try:
        generate_excel_report(result_dict, str(excel_path))
        cleanup_old_reports()
    except Exception as e:
        result_dict["excel_error"] = str(e)
        log_event(f"Excel failed for {name}: {e}", "error")

    return result_dict


# ── Scan orchestrator ───────────────────────────────────────────────────────────

def scan_and_validate_all():
    if not scan_lock.acquire(blocking=False):
        log_event("Scan already running — skipping", "warn")
        return

    scan_status.update({
        "scanning": True, "error": None,
        "current_file": None, "total_files": 0, "completed_files": 0,
    })
    try:
        pairs       = discover_pairs()
        valid_pairs = [p for p in pairs if p["has_pair"]]
        scan_status["total_files"] = len(valid_pairs)

        for pair in pairs:
            name = pair["name"]

            if not pair["has_pair"]:
                if file_states.get(name, {}).get("state") != "unmatched":
                    side  = "source" if pair["source_path"] else "target"
                    other = "target" if side == "source" else "source"
                    log_event(
                        f"{name}: found in {side} only — "
                        f"waiting for {other} to pair",
                        "warn",
                    )
                    file_states[name] = {
                        "state": "unmatched",
                        "detected_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "source_file": pair["source_file"],
                        "target_file": pair["target_file"],
                    }
                continue

            last_mtime = pair["mtime"]
            existing   = results_store.get(name)
            prev_state = file_states.get(name, {})

            if not existing:
                log_event(
                    f"{name}: new pair [{pair.get('match_type','auto')}] — "
                    f"{pair['source_file']} ↔ {pair['target_file']}",
                    "info",
                )
            elif prev_state.get("_mtime") != last_mtime:
                log_event(f"{name}: file changed — re-validating", "info")
            else:
                scan_status["completed_files"] += 1
                continue

            scan_status["current_file"] = name
            file_states[name] = {
                "state": "validating",
                "detected_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "source_file": pair["source_file"],
                "target_file": pair["target_file"],
                "_mtime": last_mtime,
            }

            try:
                result = run_validation(
                    name, pair["source_path"], pair["target_path"]
                )
                result["_mtime"] = last_mtime
                results_store[name] = result

                file_states[name] = {
                    "state":         "done",
                    "detected_at":   file_states[name]["detected_at"],
                    "validated_at":  result["run_at"],
                    "source_file":   pair["source_file"],
                    "target_file":   pair["target_file"],
                    "_mtime":        last_mtime,
                    "status":        result["status"],
                    "field_status":  result["field_status"],
                    "record_status": result["record_status"],
                    "message":       result["business_message"],
                }

                level = ("success" if result["status"] == "PASS"
                         else "warn" if result["status"] == "WARNING"
                         else "error")
                log_event(
                    f"{name}: {result['status']} — {result['business_message']} | "
                    f"Matched: {result['records_matched']:,} | "
                    f"Src only: {result['records_only_in_source']:,} | "
                    f"Tgt only: {result['records_only_in_target']:,}",
                    level,
                )
            except Exception as e:
                file_states[name]["state"] = "error"
                file_states[name]["error"] = str(e)
                scan_status["error"]       = str(e)
                log_event(f"{name}: ERROR — {e}", "error")
            finally:
                scan_status["completed_files"] += 1

        scan_status["last_scan"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    except Exception as e:
        scan_status["error"] = str(e)
        log_event(f"Scan error: {e}", "error")
    finally:
        scan_status["scanning"]     = False
        scan_status["current_file"] = None
        scan_lock.release()


def background_watcher(interval=60):
    while True:
        scan_and_validate_all()
        time.sleep(interval)


# ── Routes ──────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("dashboard.html")


@app.route("/api/scan", methods=["POST"])
def api_scan():
    threading.Thread(target=scan_and_validate_all, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/status")
def api_status():
    pairs = discover_pairs()
    cfg   = load_config()
    s, t  = get_dirs()
    sel   = cfg.get("selected_fields", [])
    tmpl  = cfg.get("active_template", "")
    return jsonify({
        "last_scan":       scan_status["last_scan"],
        "scanning":        scan_status["scanning"],
        "error":           scan_status["error"],
        "current_file":    scan_status["current_file"],
        "total_files":     scan_status["total_files"],
        "completed_files": scan_status["completed_files"],
        "source_dir":      str(s),
        "target_dir":      str(t),
        "pairs":           pairs,
        "file_states":     file_states,
        "total_tables":    len([p for p in pairs if p["has_pair"]]),
        "unmatched":       len([p for p in pairs if not p["has_pair"]]),
        "pass_threshold":  cfg.get("pass_threshold", 100.0),
        "selected_fields": sel,
        "active_template": tmpl,
        "validation_mode": (
            f"template:{tmpl}" if tmpl else
            "selected_fields"  if sel  else
            "all_fields"
        ),
    })


@app.route("/api/results")
def api_results():
    return jsonify(list(results_store.values()))


@app.route("/api/results/<name>")
def api_result_detail(name):
    r = results_store.get(name.upper())
    return jsonify(r) if r else (jsonify({"error": "Not found"}), 404)


@app.route("/api/activity")
def api_activity():
    return jsonify(list(reversed(activity_log)))


# ── Upload ──────────────────────────────────────────────────────────────────────

@app.route("/api/upload/source", methods=["POST"])
def upload_source():
    return _handle_upload(request, get_dirs()[0], "source")


@app.route("/api/upload/target", methods=["POST"])
def upload_target():
    return _handle_upload(request, get_dirs()[1], "target")


def _handle_upload(req, dest_dir, side):
    if "file" not in req.files:
        return jsonify({"error": "No file part"}), 400
    saved, errors = [], []
    for f in req.files.getlist("file"):
        if not f.filename:
            continue
        save_name = secure_filename(f.filename)
        if Path(save_name).suffix.lower() not in SUPPORTED_EXT:
            errors.append(f"Unsupported type: {save_name}")
            continue
        f.save(str(dest_dir / save_name))
        log_event(f"Uploaded to {side}: {save_name}", "info")
        saved.append(save_name)
    if saved:
        threading.Thread(target=scan_and_validate_all, daemon=True).start()
    if errors and not saved:
        return jsonify({"error": "; ".join(errors)}), 400

    # ── Read headers immediately from uploaded file so dashboard can
    # populate the field selector without waiting for a scan ──────────────────
    custom   = _get_custom_labels()
    headers  = {}
    for fname in saved:
        fpath = dest_dir / fname
        try:
            if side == "source":
                cols, _ = _read_file_headers(str(fpath))
            else:
                _, cols = _read_file_headers(None, str(fpath))
            cols = cols or []
            headers[fname] = {
                "columns": cols,
                "labels":  {c: get_label(c, custom) for c in cols},
                "count":   len(cols),
            }
            log_event(
                f"Headers read from {side}/{fname}: "
                f"{len(cols)} columns — {', '.join(cols[:6])}"
                + (" …" if len(cols) > 6 else ""),
                "info",
            )
        except Exception as e:
            headers[fname] = {"error": str(e), "columns": [], "count": 0}
            log_event(f"Could not read headers from {fname}: {e}", "warn")

    return jsonify({"ok": True, "saved": saved, "warnings": errors,
                    "side": side, "headers": headers})


@app.route("/api/upload/labels", methods=["POST"])
def upload_labels():
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "No filename"}), 400
    f.save(str(LABELS_FILE))
    log_event(f"Custom labels uploaded: {secure_filename(f.filename)}", "info")
    results_store.clear()
    for n in file_states:
        if file_states[n].get("state") == "done":
            file_states[n]["state"] = "changed"
    threading.Thread(target=scan_and_validate_all, daemon=True).start()
    return jsonify({"ok": True})


# ── Template routes ─────────────────────────────────────────────────────────────

@app.route("/api/templates", methods=["GET"])
def api_templates_list():
    return jsonify(_list_templates())


@app.route("/api/templates/upload", methods=["POST"])
def api_template_upload():
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "No filename"}), 400
    save_name = secure_filename(f.filename)
    if Path(save_name).suffix.lower() not in TEMPLATE_EXT:
        return jsonify({"error": "Use CSV, XLSX, or TXT"}), 400
    dest = TEMPLATES_DIR / save_name
    f.save(str(dest))
    fields = _read_template_fields(dest)
    log_event(
        f"Template uploaded: {save_name} ({len(fields)} fields): "
        f"{', '.join(fields[:6])}" + (" …" if len(fields) > 6 else ""),
        "info",
    )
    return jsonify({
        "ok": True, "filename": save_name,
        "field_count": len(fields), "fields": fields,
    })


@app.route("/api/templates/<filename>", methods=["DELETE"])
def api_template_delete(filename):
    safe = secure_filename(filename)
    path = TEMPLATES_DIR / safe
    if not path.exists():
        return jsonify({"error": "Not found"}), 404
    path.unlink()
    cfg = load_config()
    if cfg.get("active_template") == safe:
        cfg["active_template"] = ""
        save_config(cfg)
    log_event(f"Template deleted: {safe}", "info")
    return jsonify({"ok": True})


@app.route("/api/templates/activate", methods=["POST"])
def api_template_activate():
    data     = request.get_json(force=True)
    filename = data.get("filename", "").strip()

    if filename:
        safe = secure_filename(filename)
        path = TEMPLATES_DIR / safe
        if not path.exists():
            return jsonify({"error": f"Template not found: {safe}"}), 404
        fields = _read_template_fields(path)
        cfg    = load_config()
        cfg["active_template"] = safe
        save_config(cfg)
        results_store.clear()
        for n in file_states:
            if file_states[n].get("state") == "done":
                file_states[n]["state"] = "changed"
        threading.Thread(target=scan_and_validate_all, daemon=True).start()
        log_event(
            f"Template activated: {safe} — validating "
            f"{len(fields)} fields: "
            f"{', '.join(fields[:6])}" + (" …" if len(fields) > 6 else ""),
            "info",
        )
        return jsonify({
            "ok": True, "active_template": safe,
            "field_count": len(fields), "fields": fields,
        })

    # Deactivate
    cfg = load_config()
    cfg["active_template"] = ""
    save_config(cfg)
    results_store.clear()
    for n in file_states:
        if file_states[n].get("state") == "done":
            file_states[n]["state"] = "changed"
    threading.Thread(target=scan_and_validate_all, daemon=True).start()
    log_event("Template deactivated — validating all fields", "info")
    return jsonify({"ok": True, "active_template": ""})


@app.route("/api/templates/sample", methods=["GET"])
def api_template_sample():
    """Download a sample field-selection template CSV."""
    lines = [
        "FIELD_NAME",
        "# One SAP field name per row. Lines starting with # are ignored.",
        "# You can use SAP 4.7 names OR S/4HANA names — both work.",
        "# Example for Customer:",
        "KUNNR",
        "NAME1",
        "KTOKD",
        "LAND1",
        "STRAS",
        "ORT01",
        "PSTLZ",
        "REGIO",
        "ZTERM",
        "WAERS",
        "TELF1",
        "SPRAS",
        "ERDAT",
    ]
    return Response(
        "\n".join(lines).encode("utf-8"),
        mimetype="text/csv",
        headers={"Content-Disposition":
                 "attachment; filename=sample_field_template.csv"},
    )


# ── Config ──────────────────────────────────────────────────────────────────────

@app.route("/api/config", methods=["GET"])
def api_get_config():
    cfg    = load_config()
    custom = _get_custom_labels()
    sel    = cfg.get("selected_fields", [])
    sel_set = set(sel)

    # ── Build available_fields from TARGET file columns (authoritative) ──────
    # Target file has ALL expected S/4HANA fields — use those as the field list.
    # Source columns are shown as secondary (S badge) so user can see coverage.
    src_dir, tgt_dir = get_dirs()
    available = []

    try:
        src_files = sorted([f for f in src_dir.iterdir()
                            if f.suffix.lower() in SUPPORTED_EXT],
                           key=lambda f: f.stat().st_mtime, reverse=True)
        tgt_files = sorted([f for f in tgt_dir.iterdir()
                            if f.suffix.lower() in SUPPORTED_EXT],
                           key=lambda f: f.stat().st_mtime, reverse=True)

        src_path = str(src_files[0]) if src_files else None
        tgt_path = str(tgt_files[0]) if tgt_files else None

        if tgt_path or src_path:
            src_cols, tgt_cols = _read_file_headers(src_path or "", tgt_path or "")
            src_set = set(src_cols)
            tgt_set = set(tgt_cols)

            # Target columns are the master list
            # Mark each as: in_target=True always, in_source=True if also in source
            for col in sorted(tgt_set):
                available.append({
                    "field":     col,
                    "label":     get_label(col, custom),
                    "in_source": col in src_set,
                    "in_target": True,
                    "common":    col in src_set,
                    "selected":  not sel_set or col in sel_set,
                })
            # Add source-only columns (in source but not target)
            for col in sorted(src_set - tgt_set):
                available.append({
                    "field":     col,
                    "label":     get_label(col, custom),
                    "in_source": True,
                    "in_target": False,
                    "common":    False,
                    "selected":  False,
                })
    except Exception as e:
        log_event(f"Config: could not read file headers from disk — {e}", "warn")

    # Fallback to last scan result if disk read produced nothing
    if not available and results_store:
        first     = next(iter(results_store.values()))
        available = first.get("available_fields", [
            {"field": fr["field"], "label": get_label(fr["field"], custom),
             "in_source": True, "in_target": True, "common": True,
             "selected":  not sel_set or fr["field"] in sel_set}
            for fr in first.get("field_results", [])
        ])

    # Also include which files are currently on disk so the UI can show them
    src_dir2, tgt_dir2 = get_dirs()
    src_files_list = sorted([f.name for f in src_dir2.iterdir()
                              if f.suffix.lower() in SUPPORTED_EXT])
    tgt_files_list = sorted([f.name for f in tgt_dir2.iterdir()
                              if f.suffix.lower() in SUPPORTED_EXT])

    return jsonify({
        "source_dir":         cfg.get("source_dir",      DEFAULT_CONFIG["source_dir"]),
        "target_dir":         cfg.get("target_dir",      DEFAULT_CONFIG["target_dir"]),
        "pass_threshold":     cfg.get("pass_threshold",  100.0),
        "selected_fields":    sel,
        "active_template":    cfg.get("active_template", ""),
        "available_fields":   available,
        "source_files":       src_files_list,
        "target_files":       tgt_files_list,
        "labels_file_exists": LABELS_FILE.exists(),
        "labels_file":        str(LABELS_FILE) if LABELS_FILE.exists() else None,
    })


@app.route("/api/config", methods=["POST"])
def api_set_config():
    data    = request.get_json(force=True)
    cfg     = load_config()
    changed = False

    for key in ("source_dir", "target_dir"):
        if key in data and str(data[key]).strip():
            np = str(Path(str(data[key]).strip()))
            if np != cfg.get(key):
                cfg[key] = np
                changed  = True

    if "pass_threshold" in data:
        thr = float(data["pass_threshold"])
        if thr != cfg.get("pass_threshold"):
            cfg["pass_threshold"] = thr
            changed = True
            log_event(f"Pass threshold → {thr}%", "info")

    if "selected_fields" in data:
        sel = [str(f).strip().upper() for f in data["selected_fields"]
               if str(f).strip()]
        if sel != cfg.get("selected_fields", []):
            cfg["selected_fields"] = sel
            changed = True
            log_event(
                f"Field selection: {len(sel)} fields" if sel
                else "Field selection: all fields",
                "info",
            )

    if changed:
        save_config(cfg)
        results_store.clear()
        for n in file_states:
            if file_states[n].get("state") == "done":
                file_states[n]["state"] = "changed"
        threading.Thread(target=scan_and_validate_all, daemon=True).start()

    return jsonify({"ok": True, "config": cfg})


# ── Field preview ────────────────────────────────────────────────────────────────

@app.route("/api/fields/from-files", methods=["POST"])
def api_fields_from_files():
    """
    Read headers from specific files already on disk by filename.
    Body: {"source_file": "customers.csv", "target_file": "Export_Data.csv"}
    Called when user picks a specific file pair to load fields from.
    Returns full field list with labels immediately — no scan needed.
    """
    data     = request.get_json(force=True)
    src_name = data.get("source_file", "").strip()
    tgt_name = data.get("target_file", "").strip()
    custom   = _get_custom_labels()
    sel      = load_config().get("selected_fields", [])
    sel_set  = set(sel)

    src_dir, tgt_dir = get_dirs()
    src_path = str(src_dir / src_name) if src_name else None
    tgt_path = str(tgt_dir / tgt_name) if tgt_name else None

    errors = {}
    src_cols, tgt_cols = [], []

    if src_path and Path(src_path).exists():
        try:
            src_cols, _ = _read_file_headers(src_path)
        except Exception as e:
            errors["source"] = str(e)
    elif src_name:
        errors["source"] = f"File not found: {src_name}"

    if tgt_path and Path(tgt_path).exists():
        try:
            _, tgt_cols = _read_file_headers(None, tgt_path)
        except Exception as e:
            errors["target"] = str(e)
    elif tgt_name:
        errors["target"] = f"File not found: {tgt_name}"

    src_set  = set(src_cols)
    tgt_set  = set(tgt_cols)

    fields = []
    # TARGET columns are the master list — show all target fields first
    for col in sorted(tgt_set):
        fields.append({
            "field":     col,
            "label":     get_label(col, custom),
            "in_source": col in src_set,
            "in_target": True,
            "common":    col in src_set,
            "selected":  not sel_set or col in sel_set,
        })
    # Source-only columns (in source but not in target)
    for col in sorted(src_set - tgt_set):
        fields.append({
            "field":     col,
            "label":     get_label(col, custom),
            "in_source": True,
            "in_target": False,
            "common":    False,
            "selected":  False,
        })

    common   = len(src_set & tgt_set)
    src_only = len(src_set - tgt_set)
    tgt_only = len(tgt_set - src_set)

    log_event(
        f"Fields from files: "
        f"src={src_name}({len(src_cols)}) "
        f"tgt={tgt_name}({len(tgt_cols)}) "
        f"→ {common} common, {src_only} src-only, {tgt_only} tgt-only",
        "info",
    )

    return jsonify({
        "fields":      fields,
        "src_count":   len(src_cols),
        "tgt_count":   len(tgt_cols),
        "common":      common,
        "src_only":    src_only,
        "tgt_only":    tgt_only,
        "errors":      errors,
        "source_file": src_name,
        "target_file": tgt_name,
    })


@app.route("/api/fields/preview", methods=["POST"])
def api_fields_preview():
    data      = request.get_json(force=True)
    src_path  = data.get("source_path", "").strip()
    tgt_path  = data.get("target_path", "").strip()
    src_delim = data.get("source_delimiter", ",")
    tgt_delim = data.get("target_delimiter", ",")
    custom    = _get_custom_labels()

    def read_headers(path, delim):
        p = Path(path)
        if not p.exists():
            return None, f"File not found: {path}"
        try:
            if p.suffix.lower() in (".xlsx", ".xls"):
                import openpyxl
                wb   = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
                ws   = wb.active
                cols = [str(c.value).strip().upper()
                        for c in next(ws.iter_rows(max_row=1)) if c.value]
                wb.close()
            else:
                import csv
                with open(str(p), encoding="utf-8-sig") as f:
                    cols = [c.strip().upper()
                            for c in next(csv.reader(f, delimiter=delim))]
            return cols, None
        except Exception as e:
            return None, str(e)

    src_cols, src_err = read_headers(src_path, src_delim) if src_path else ([], None)
    tgt_cols, tgt_err = read_headers(tgt_path, tgt_delim) if tgt_path else ([], None)

    errors = {}
    if src_err: errors["source"] = src_err
    if tgt_err: errors["target"] = tgt_err

    src_set  = set(src_cols or [])
    tgt_set  = set(tgt_cols or [])
    common   = sorted(src_set & tgt_set)
    src_only = sorted(src_set - tgt_set)
    tgt_only = sorted(tgt_set - src_set)

    cfg          = load_config()
    selected_set = set(cfg.get("selected_fields", []))

    fields = []
    for col in common:
        fields.append({
            "field": col, "label": get_label(col, custom),
            "in_source": True, "in_target": True, "common": True,
            "selected": not selected_set or col in selected_set,
        })
    for col in src_only:
        fields.append({
            "field": col, "label": get_label(col, custom),
            "in_source": True, "in_target": False, "common": False,
            "selected": False,
        })
    for col in tgt_only:
        fields.append({
            "field": col, "label": get_label(col, custom),
            "in_source": False, "in_target": True, "common": False,
            "selected": False,
        })

    return jsonify({
        "fields":    fields,
        "src_count": len(src_cols or []),
        "tgt_count": len(tgt_cols or []),
        "common":    len(common),
        "src_only":  len(src_only),
        "tgt_only":  len(tgt_only),
        "errors":    errors,
    })


# ── Files & pairs ────────────────────────────────────────────────────────────────

@app.route("/api/files/list")
def api_files_list():
    src, tgt = get_available_files()
    return jsonify({
        "source_files": [f.name for f in src],
        "target_files": [f.name for f in tgt],
    })


@app.route("/api/pairs", methods=["GET"])
def api_pairs_get():
    return jsonify(load_config().get("manual_pairs", []))


@app.route("/api/pairs", methods=["POST"])
def api_pairs_save():
    data  = request.get_json(force=True)
    seen  = set()
    clean = []
    for p in data.get("pairs", []):
        name = str(p.get("name", "")).strip().upper()
        sf   = str(p.get("source_file", "")).strip()
        tf   = str(p.get("target_file", "")).strip()
        if not name or not sf or not tf or name in seen:
            continue
        seen.add(name)
        clean.append({"name": name, "source_file": sf, "target_file": tf})
    cfg = load_config()
    cfg["manual_pairs"] = clean
    save_config(cfg)
    results_store.clear()
    for n in list(file_states):
        if file_states[n].get("state") == "done":
            file_states[n]["state"] = "changed"
    log_event(f"Manual pairs updated: {len(clean)} pair(s)", "info")
    threading.Thread(target=scan_and_validate_all, daemon=True).start()
    return jsonify({"ok": True, "saved": len(clean)})


@app.route("/api/pairs/<name>", methods=["DELETE"])
def api_pairs_delete(name):
    cfg    = load_config()
    pairs  = cfg.get("manual_pairs", [])
    before = len(pairs)
    pairs  = [p for p in pairs if p["name"].upper() != name.upper()]
    cfg["manual_pairs"] = pairs
    save_config(cfg)
    removed = before - len(pairs)
    if removed:
        results_store.pop(name.upper(), None)
        file_states.pop(name.upper(), None)
        log_event(f"Manual pair removed: {name}", "info")
    return jsonify({"ok": True, "removed": removed})


# ── Labels / reports / downloads ─────────────────────────────────────────────────

@app.route("/api/labels/sample")
def api_labels_sample():
    lines = ["FIELD_NAME,FRIENDLY_LABEL"] + [
        f"{k},{v}" for k, v in list(SAP_FIELD_LABELS.items())[:25]
    ]
    return Response(
        "\n".join(lines).encode("utf-8"),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=sample_labels.csv"},
    )


@app.route("/api/objects")
def api_objects():
    custom = _get_custom_labels()
    return jsonify([
        {
            "key":            k,
            "description":    v.get("description", k),
            "join_key":       v.get("join_key", ""),
            "join_key_label": get_label(v.get("join_key", ""), custom),
            "key_fields":     [
                {"field": f, "label": get_label(f, custom)}
                for f in v.get("key_fields", [])
            ],
        }
        for k, v in SAP_OBJECT_CONFIG.items()
    ])


@app.route("/api/download/<name>")
def api_download(name):
    r = results_store.get(name.upper())
    if not r:
        return jsonify({"error": "Not found"}), 404
    path = REPORTS_DIR / r.get("excel_file", "")
    if not path.exists():
        return jsonify({"error": "Report file missing"}), 404
    return send_file(
        str(path), as_attachment=True, download_name=path.name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/download-file/<filename>")
def api_download_file(filename):
    path = REPORTS_DIR / filename
    if not path.exists() or not filename.endswith(".xlsx"):
        return jsonify({"error": "Not found"}), 404
    return send_file(
        str(path), as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/reports")
def api_reports():
    files = sorted(
        REPORTS_DIR.glob("*.xlsx"),
        key=lambda f: f.stat().st_mtime,
        reverse=True,
    )
    return jsonify([
        {
            "filename": f.name,
            "size_kb":  round(f.stat().st_size / 1024, 1),
            "modified": datetime.fromtimestamp(
                f.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
        }
        for f in files
    ])


@app.route("/api/folders")
def api_folders():
    s, t = get_dirs()
    return jsonify({
        "source_dir":  str(s),
        "target_dir":  str(t),
        "reports_dir": str(REPORTS_DIR),
    })


@app.route("/api/clear-results", methods=["POST"])
def api_clear_results():
    results_store.clear()
    file_states.clear()
    activity_log.clear()
    scan_status.update({
        "last_scan": None, "scanning": False, "error": None,
        "current_file": None, "total_files": 0, "completed_files": 0,
    })
    log_event("Results cleared", "info")
    return jsonify({"ok": True})


# ── Entry point ─────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    s, t = get_dirs()
    cfg  = load_config()
    print("\n  ╔══════════════════════════════════════╗")
    print("  ║  Genpact SAP Migration Validator V4  ║")
    print("  ╚══════════════════════════════════════╝")
    print(f"  Source dir     → {s}")
    print(f"  Target dir     → {t}")
    print(f"  Templates      → {TEMPLATES_DIR}")
    print(f"  Reports        → {REPORTS_DIR}")
    print(f"  Pass threshold → {cfg.get('pass_threshold', 100)}%")
    if cfg.get("active_template"):
        print(f"  Active template→ {cfg['active_template']}")
    print("  Open           → http://localhost:5000\n")
    threading.Thread(target=scan_and_validate_all, daemon=True).start()
    # threading.Thread(target=background_watcher, args=(60,), daemon=True).start()
    app.run(debug=False, port=5000, use_reloader=False)
