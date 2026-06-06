"""
SAP Migration Post-Load Validator — Excel Report Generator V4
Accepts a ValidationResult dataclass or the result_dict from app.py.
"""

from __future__ import annotations
from datetime import datetime
from pathlib import Path


def _coerce(result) -> dict:
    if isinstance(result, dict):
        return result
    ss      = result.summary_stats
    mapping = result.mapping
    thr     = mapping.pass_threshold if mapping else 100.0
    sel     = mapping.selected_fields if mapping else []
    rows    = []
    for fr in result.field_results:
        rows.append({
            "field":          fr.field_source,
            "field_label":    fr.field_label,
            "field_target":   fr.field_target,
            "type":           "numeric" if fr.is_numeric else "string",
            "tolerance":      fr.tolerance_used,
            "total":          fr.total_records,
            "matched":        fr.matched,
            "mismatched":     fr.mismatched,
            "miss_source":    fr.missing_in_source,
            "miss_target":    fr.missing_in_target,
            "match_pct":      fr.match_pct,
            "pass_threshold": fr.pass_threshold,
            "status":         fr.status,
            "mismatches":     fr.mismatch_details,
            "display_name":   fr.field_label,
            "is_cross_mapped": fr.field_source != fr.field_target,
        })
    mp = None
    if mapping:
        mp = {
            "join_key":           mapping.join_key,
            "join_key_label":     mapping.join_key_label,
            "numeric_fields":     mapping.numeric_fields,
            "tolerance_map":      mapping.tolerance_map,
            "source_only_fields": mapping.source_only_fields,
            "target_only_fields": mapping.target_only_fields,
            "selected_fields":    sel,
            "pass_threshold":     thr,
        }
    return {
        "name":                   Path(result.source_file).stem.upper(),
        "status":                 result.overall_status,
        "source_file":            result.source_file,
        "target_file":            result.target_file,
        "total_source_records":   result.total_source_records,
        "total_target_records":   result.total_target_records,
        "records_matched":        result.records_matched,
        "records_only_in_source": result.records_only_in_source,
        "records_only_in_target": result.records_only_in_target,
        "fields_passed":          ss["fields_passed"],
        "fields_failed":          ss["fields_failed"],
        "total_fields":           ss["total_fields_validated"],
        "pass_rate_pct":          ss["pass_rate_pct"],
        "errors":                 result.errors,
        "mapping":                mp,
        "field_results":          rows,
        "run_at":                 datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def generate_excel_report(result, output_path: str) -> str:
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("pip install openpyxl")

    r        = _coerce(result)
    mapping  = r.get("mapping") or {}
    pass_thr = mapping.get("pass_threshold", 100.0)
    sel      = mapping.get("selected_fields", [])

    C_NAVY  = "FF1B3A57"; C_WHITE  = "FFFFFFFF"
    C_GREEN = "FF00AA44"; C_RED    = "FFCC2200"
    C_AMBER = "FFDD8800"; C_DARK   = "FF333333"
    C_LG    = "FFE6F4EA"; C_LR     = "FFFCE8E6"; C_LGREY = "FFF5F5F5"

    def fill(c):  return PatternFill("solid", fgColor=c)
    def bdr():
        s = Side(style="thin", color="FFCCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)
    def hcell(ws, row, col, val, bg=C_NAVY):
        c = ws.cell(row, col, val)
        c.fill = fill(bg); c.font = Font(bold=True, color=C_WHITE, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr(); return c

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = f"SAP Post-Load Validation — {r['name']}"
    c.font  = Font(bold=True, size=16, color=C_WHITE)
    c.fill  = fill(C_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    oc = C_GREEN if r["status"] == "PASS" else C_RED
    meta = [
        ("Run Date",       r.get("run_at", "")),
        ("Source File",    r["source_file"]),
        ("Target File",    r["target_file"]),
        ("Pass Threshold", f">= {pass_thr}%"),
        ("Fields Scope",   f"{len(sel)} selected" if sel else "All common fields"),
        ("Overall Status", r["status"]),
    ]
    for i, (k, v) in enumerate(meta, start=3):
        ws.cell(i, 1, k).font = Font(bold=True, color="FF444444")
        cell = ws.cell(i, 2, v)
        if k == "Overall Status": cell.font = Font(bold=True, color=oc, size=12)
        if k == "Pass Threshold": cell.font = Font(bold=True, color=C_AMBER)

    ws.cell(11, 1, "Record Counts").font = Font(bold=True, size=11)
    for i, (k, v) in enumerate([
        ("Source Records", r["total_source_records"]),
        ("Target Records", r["total_target_records"]),
        ("Keys Matched",   r["records_matched"]),
        ("Source Only",    r["records_only_in_source"]),
        ("Target Only",    r["records_only_in_target"]),
    ], start=12):
        ws.cell(i, 1, k)
        cell = ws.cell(i, 2, v)
        if k in ("Source Only", "Target Only") and v > 0:
            cell.font = Font(bold=True, color=C_RED)

    ws.cell(11, 4, "Validation Stats").font = Font(bold=True, size=11)
    for i, (k, v) in enumerate([
        ("Fields Validated", r["total_fields"]),
        ("Fields Passed",    r["fields_passed"]),
        ("Fields Failed",    r["fields_failed"]),
        ("Pass Rate",        f"{r['pass_rate_pct']}%"),
    ], start=12):
        ws.cell(i, 4, k)
        cell = ws.cell(i, 5, v)
        if k == "Fields Failed" and isinstance(v, int) and v > 0:
            cell.font = Font(bold=True, color=C_RED)
        if k == "Pass Rate":
            cell.font = Font(bold=True, color=C_GREEN if r["fields_failed"] == 0 else C_RED)

    if mapping:
        ws.cell(11, 7, "Auto-Detected").font = Font(bold=True, size=11)
        ws.cell(12, 7, "Join Key")
        ws.cell(12, 8, f"{mapping.get('join_key_label', '')}  ({mapping.get('join_key', '')})")
        ws.cell(13, 7, "Numeric Fields")
        ws.cell(13, 8, ", ".join(mapping.get("numeric_fields", [])) or "none")

    ws.cell(19, 1, "Field-Level Results").font = Font(bold=True, size=11)
    ws.cell(20, 1, f"Pass threshold: >= {pass_thr}%").font = Font(color=C_AMBER, size=10, italic=True)

    hdrs = ["Field Label","Source Field","Target Field","Map Method","Type","Tolerance",
            "Total","Matched","Mismatched","Miss-Src","Miss-Tgt","Match %","Threshold","Status"]
    for col, h in enumerate(hdrs, 1):
        hcell(ws, 22, col, h)

    for ri, fr in enumerate(r["field_results"], start=23):
        thr    = fr.get("pass_threshold", pass_thr)
        pct    = fr["match_pct"]
        status = fr["status"]
        bg     = C_LG if status == "PASS" else C_LR
        tol    = f"+-{fr['tolerance']}" if fr.get("tolerance") is not None else "—"
        label  = fr.get("display_name") or fr.get("field_label") or fr["field"]
        method = fr.get("mapping_method", "exact")
        vals   = [label, fr["field"], fr.get("field_target", fr["field"]), method,
                  fr.get("type",""), tol, fr["total"], fr["matched"], fr["mismatched"],
                  fr["miss_source"], fr["miss_target"], f"{pct}%", f">= {thr}%", status]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(ri, col, val)
            cell.fill = fill(bg); cell.border = bdr()
            cell.alignment = Alignment(horizontal="center" if col > 2 else "left", vertical="center")
            if col == len(vals): cell.font = Font(bold=True, color=C_GREEN if status == "PASS" else C_RED)
            if col == 12: cell.font = Font(bold=True, color=C_GREEN if pct >= thr else C_RED)

    col_widths = [26, 14, 14, 14, 8, 10, 8, 10, 12, 10, 10, 10, 12, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for fr in r["field_results"]:
        if fr["status"] != "FAIL" or not fr.get("mismatches"):
            continue
        label  = fr.get("display_name") or fr.get("field_label") or fr["field"]
        safe   = label[:22].replace("/", "_").replace("\\", "_")
        ws2    = wb.create_sheet(title=f"FAIL_{safe}")
        ws2.sheet_view.showGridLines = False

        ws2.merge_cells("A1:F1")
        c2 = ws2["A1"]
        src_tgt = fr["field"] if fr["field"] == fr.get("field_target", fr["field"]) \
                  else f"{fr['field']} \u2192 {fr.get('field_target', '')}"
        c2.value = f"Mismatches — {label}  ({src_tgt})"
        c2.font  = Font(bold=True, size=13, color=C_WHITE)
        c2.fill  = fill(C_RED)
        c2.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 28

        ws2.merge_cells("A2:F2")
        ws2["A2"].value = (f"Match: {fr['match_pct']}%  |  Threshold: >= {thr}%  |  "
                           f"Total: {fr['total']}  |  Matched: {fr['matched']}")
        ws2["A2"].font  = Font(size=10, color="FF555555")
        ws2["A2"].fill  = fill(C_LR)
        ws2["A2"].alignment = Alignment(horizontal="center")

        for col, h in enumerate(["Key", "Source Value", "Target Value", "Issue", "Source Field", "Target Field"], 1):
            hcell(ws2, 4, col, h, bg=C_DARK)

        for ri2, rec in enumerate(fr["mismatches"], start=5):
            bg2 = C_LR if ri2 % 2 == 0 else C_LGREY
            for ci, v in enumerate([
                rec.get("material",""), rec.get("source_value",""),
                rec.get("target_value",""), rec.get("issue",""),
                fr["field"], fr.get("field_target", fr["field"])
            ], 1):
                cell = ws2.cell(ri2, ci, v)
                cell.fill = fill(bg2); cell.border = bdr()
                if ci == 2: cell.font = Font(color=C_RED)
                if ci == 3: cell.font = Font(color=C_GREEN)

        for col, w in zip("ABCDEF", [24, 30, 30, 36, 14, 14]):
            ws2.column_dimensions[col].width = w

    wb.save(output_path)
    return output_path
